#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工程造价清单记忆库 — list_memory.py
自动记录每次审核的清单数据，积累为审核知识库
"""

import os
import json
import sqlite3
from datetime import datetime
from openpyxl import load_workbook

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "zaojiatong.db")

# ═══════════════════════════════════════════════
# 清单记忆表
# ═══════════════════════════════════════════════

def init_memory_tables():
    """创建记忆相关的数据库表"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # 已审清单项目表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audited_projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_name TEXT NOT NULL,
            building_type TEXT DEFAULT '工业建筑',
            area REAL DEFAULT 0,
            total_price REAL DEFAULT 0,
            file_path TEXT DEFAULT '',
            audit_date TEXT DEFAULT (datetime('now','localtime')),
            notes TEXT DEFAULT ''
        )
    """)

    # 已审清单子目明细表（记忆每条审核过的单价）
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audited_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            item_name TEXT NOT NULL,
            item_desc TEXT DEFAULT '',
            unit TEXT DEFAULT '',
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            total_price REAL DEFAULT 0,
            price_rating TEXT DEFAULT 'normal',
            market_price_lo REAL DEFAULT 0,
            market_price_hi REAL DEFAULT 0,
            deviation_pct REAL DEFAULT 0,
            audit_date TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (project_id) REFERENCES audited_projects(id)
        )
    """)

    # 材料价格记忆库（每次审核时记录的价格，逐步积累）
    cur.execute("""
        CREATE TABLE IF NOT EXISTS material_price_memory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            material_name TEXT NOT NULL,
            spec TEXT DEFAULT '',
            unit TEXT DEFAULT '',
            price REAL DEFAULT 0,
            source TEXT DEFAULT '',
            project_name TEXT DEFAULT '',
            recorded_date TEXT DEFAULT (datetime('now','localtime')),
            UNIQUE(material_name, spec)
        )
    """)

    # 单价记忆库（按子目名称累计的历史平均价）
    cur.execute("""
        CREATE TABLE IF NOT EXISTS price_memory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_name TEXT NOT NULL UNIQUE,
            avg_price REAL DEFAULT 0,
            min_price REAL DEFAULT 0,
            max_price REAL DEFAULT 0,
            sample_count INTEGER DEFAULT 1,
            last_updated TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    conn.commit()
    conn.close()


def safe_float(val):
    try: return float(str(val).replace(",", "").replace("元", "").replace(" ", ""))
    except: return 0.0


def safe_str(val):
    return str(val).strip() if val else ""


# ═══════════════════════════════════════════════
# 保存清单到记忆库
# ═══════════════════════════════════════════════

def remember_list(filepath: str, project_name: str = "", building_type: str = "工业建筑", area: float = 0) -> int:
    """
    将一份已标价清单存入记忆库
    返回 project_id
    """
    init_memory_tables()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    if not project_name:
        project_name = os.path.splitext(os.path.basename(filepath))[0]

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # 查找表头
    header_idx = 0
    for i, row in enumerate(rows):
        texts = [str(c).strip() for c in row if c]
        combined = " ".join(texts)
        keywords = ["项目名称", "子目名称", "清单编号", "计量单位", "综合单价"]
        if sum(1 for kw in keywords if kw in combined) >= 3:
            header_idx = i
            break

    headers = [safe_str(c) for c in rows[header_idx]]

    def col(keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h:
                    return i
        return -1

    name_col = col(["项目名称", "子目名称", "名称"])
    desc_col = col(["项目特征", "特征描述", "描述"])
    unit_col = col(["计量单位", "单位"])
    qty_col  = col(["工程数量", "工程量", "数量"])
    price_col = col(["综合单价", "单价"])

    total_price = 0
    item_count = 0

    # 创建项目记录
    cur.execute(
        "INSERT INTO audited_projects (project_name, building_type, area, total_price, file_path) VALUES (?,?,?,?,?)",
        (project_name, building_type, area, 0, filepath)
    )
    project_id = cur.lastrowid

    # 逐条保存子目
    for r in rows[header_idx + 1:]:
        vals = [safe_str(c) for c in r]
        if not any(v for v in vals):
            continue

        name = safe_str(r[name_col]) if name_col >= 0 and name_col < len(r) else ""
        if not name:
            continue

        desc = safe_str(r[desc_col]) if desc_col >= 0 and desc_col < len(r) else ""
        unit = safe_str(r[unit_col]) if unit_col >= 0 and unit_col < len(r) else ""
        qty  = safe_float(r[qty_col]) if qty_col >= 0 and qty_col < len(r) else 0
        price = safe_float(r[price_col]) if price_col >= 0 and price_col < len(r) else 0
        total = round(price * qty, 2)

        # 判定价格水平（基于内置规则）
        rating = "normal"
        from zaojia_auditor import PRICE_RANGES
        for pat, (lo, hi, u) in PRICE_RANGES.items():
            if pat in name and price > 0:
                mid = (lo + hi) / 2
                dev = abs(price - mid) / mid * 100
                if dev > 20: rating = "high"
                elif dev > 10: rating = "medium"
                break

        cur.execute(
            """INSERT INTO audited_items (project_id, item_name, item_desc, unit, quantity, unit_price, total_price, price_rating, market_price_lo, market_price_hi)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (project_id, name, desc, unit, qty, price, total, rating, 0, 0)
        )
        total_price += total
        item_count += 1

        # 更新单价记忆库
        cur.execute(
            """INSERT INTO price_memory (item_name, avg_price, min_price, max_price, sample_count)
               VALUES (?, ?, ?, ?, 1)
               ON CONFLICT(item_name) DO UPDATE SET
                 sample_count = sample_count + 1,
                 avg_price = (avg_price * sample_count + ?) / (sample_count + 1),
                 min_price = MIN(min_price, ?),
                 max_price = MAX(max_price, ?),
                 last_updated = datetime('now','localtime')""",
            (name[:60], price, price, price, price, price, price)
        )

    # 更新项目总造价
    cur.execute("UPDATE audited_projects SET total_price=? WHERE id=?", (total_price, project_id))
    conn.commit()
    conn.close()

    print(f"OK:  已记住清单: {project_name}")
    print(f"   项目ID: {project_id}")
    print(f"   子目数: {item_count} 条")
    print(f"   总造价: {total_price/10000:.2f} 万元")
    return project_id


# ═══════════════════════════════════════════════
# 查询记忆库
# ═══════════════════════════════════════════════

def recall_similar(name: str, limit: int = 5) -> list:
    """查询相似子目的历史价格记忆"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # 精确匹配
    rows = cur.execute(
        "SELECT item_name, avg_price, min_price, max_price, sample_count FROM price_memory WHERE item_name=? LIMIT 1",
        (name,)
    ).fetchall()

    if not rows:
        # 模糊匹配
        rows = cur.execute(
            "SELECT item_name, avg_price, min_price, max_price, sample_count FROM price_memory WHERE item_name LIKE ? LIMIT ?",
            (f"%{name[:8]}%", limit)
        ).fetchall()

    conn.close()
    return [{"name": r[0], "avg": round(r[1], 2), "min": round(r[2], 2),
             "max": round(r[3], 2), "samples": r[4]} for r in rows]


def get_all_projects() -> list:
    """获取所有已审项目列表"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    rows = cur.execute(
        "SELECT id, project_name, building_type, area, total_price, audit_date FROM audited_projects ORDER BY audit_date DESC"
    ).fetchall()
    conn.close()
    return [{"id": r[0], "name": r[1], "type": r[2], "area": r[3],
             "total": r[4], "date": r[5]} for r in rows]


def get_memory_stats() -> dict:
    """获取记忆库统计"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    projects = cur.execute("SELECT COUNT(*) FROM audited_projects").fetchone()[0]
    items = cur.execute("SELECT COUNT(DISTINCT item_name) FROM price_memory").fetchone()[0]
    conn.close()
    return {"projects": projects, "unique_items": items}


# ═══════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════

if __name__ == "__main__":
    import sys
    init_memory_tables()

    if len(sys.argv) < 2:
        stats = get_memory_stats()
        print(f"MEM:  清单记忆库统计")
        print(f"   已审项目: {stats['projects']} 个")
        print(f"   已记子目类型: {stats['unique_items']} 种")
        print(f"\n用法:")
        print(f"  python list_memory.py remember <文件路径> [项目名] [建筑类型] [面积]")
        print(f"  python list_memory.py recall <子目名称>")
        print(f"  python list_memory.py projects")
    elif sys.argv[1] == "remember" and len(sys.argv) > 2:
        path = sys.argv[2]
        pname = sys.argv[3] if len(sys.argv) > 3 else ""
        btype = sys.argv[4] if len(sys.argv) > 4 else "工业建筑"
        area = float(sys.argv[5]) if len(sys.argv) > 5 else 0
        remember_list(path, pname, btype, area)
    elif sys.argv[1] == "recall" and len(sys.argv) > 2:
        results = recall_similar(sys.argv[2])
        for r in results:
            print(f"ITEM:  {r['name']}")
            print(f"   平均 {r['avg']}元  范围 {r['min']}~{r['max']}  (来自{r['samples']}个项目)")
    elif sys.argv[1] == "projects":
        for p in get_all_projects():
            print(f"  [{p['id']}] {p['name']} | {p['type']} | {p['total']/10000:.1f}万元 | {p['date']}")
