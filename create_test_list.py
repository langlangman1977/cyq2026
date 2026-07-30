#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成测试用工程造价清单 Excel 文件
模拟一份包含 30 条子目的工业项目（破碎楼）工程量清单
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "工程量清单"

# 样式
header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
body_font = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center")

# 表头
headers = ["序号", "项目编码", "项目名称", "项目特征描述", "计量单位", "工程数量", "综合单价", "合价"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

# 清单数据 (名称, 描述, 单位, 数量, 单价) -- 部分故意设偏差
items = [
    ("挖一般土方", "机械开挖，场内运输1km", "m³", 850, 6.5),        # 正常
    ("土方回填", "分层夯实，每层30cm", "m³", 450, 18),           # 正常
    ("现浇混凝土基础 C30", "独立基础 含模板钢筋", "m³", 120, 780),   # 正常
    ("现浇混凝土柱 C35", "矩形柱 含模板钢筋", "m³", 68, 1350),      # 🔴偏高 ~23%
    ("现浇混凝土梁 C30", "矩形梁 含模板钢筋", "m³", 42, 1420),      # 🔴偏高 ~32%
    ("现浇混凝土板 C30", "有梁板 含模板钢筋", "m³", 95, 1100),      # 正常
    ("钢筋制作安装 HRB400", "含加工绑扎 直径12-25", "t", 85, 6850), # 🔴偏高 ~15%
    ("钢结构制作安装 Q355B", "含除锈防腐防火涂料 焊接H型钢", "t", 45, 10500), # 🔴偏高 ~12%
    ("防火涂料 薄型 2h", "钢柱钢梁 耐火极限2小时", "m²", 680, 52),   # 🔴偏高 ~35%
    ("加气混凝土砌块墙", "600x240x200 B06 含砂浆构造柱", "m³", 35, 498), # 正常
    ("内墙抹灰 1:3水泥砂浆", "15mm厚 含基层处理", "m²", 1200, 32),    # 正常
    ("外墙保温 50mm XPS板", "含网格布+抗裂砂浆", "m²", 850, 98),     # 正常
    ("屋面防水 SBS 3mm两道", "含找平层", "m²", 420, 88),          # 正常
    ("铝合金窗 断桥隔热", "5+12A+5中空玻璃", "m²", 180, 780),      # 🟡偏高 ~8%
    ("钢制防火门 甲级", "含门框门扇五金", "m²", 24, 720),          # 🔴偏高 ~25%
    ("地砖铺贴 800x800", "含找平层 水泥砂浆", "m²", 650, 62),       # 正常
    ("内墙涂料 乳胶漆", "腻子两遍+乳胶漆两遍", "m²", 2400, 42),      # 🔴偏高 ~40%
    ("外墙真石漆", "含底漆+面漆+罩光漆", "m²", 850, 72),          # 正常
    ("电气配管 PVC20", "含接线盒 暗敷", "m", 1200, 10),           # 正常
    ("电气配线 BV-2.5mm²", "管内穿线", "m", 3600, 5),           # 正常
    ("给水管道 PPR DN20", "含管件+试压", "m", 280, 38),          # 正常
    ("排水管道 PVC-U DN110", "含管件+灌水试验", "m", 160, 42),      # 正常
    ("成品钢爬梯", "含制作安装防腐", "t", 2, 18000),               # 🔴偏高 ~50% 非标件
    ("预埋铁件", "含制作安装", "t", 3.5, 8500),                  # 正常
    ("现浇混凝土设备基础 C35", "大型破碎机基础 含模板钢筋", "m³", 220, 950), # 🔴偏高 ~18%
    ("钢格栅板平台", "含制作安装防腐", "t", 8, 12000),             # 正常
    ("H型钢柱 Q355B", "含制作安装防腐防火", "t", 28, 9200),        # 正常
    ("预埋螺栓 M36", "含安装", "个", 120, 85),                  # 🔴偏高 ~30%
    ("钢筋网片 Φ6@100", "含制作安装", "m²", 450, 15),             # 正常
]

for i, (name, desc, unit, qty, price) in enumerate(items, 1):
    total = round(qty * price, 2)
    data = [i, f"010{i:03d}001001", name, desc, unit, qty, price, total]
    for j, val in enumerate(data, 1):
        c = ws.cell(row=i + 1, column=j, value=val)
        c.font = body_font
        c.border = thin_border
        c.alignment = center if j in (1, 4, 5, 6, 7, 8) else Alignment(horizontal="left", vertical="center")

# 列宽
widths = [6, 14, 30, 35, 8, 10, 12, 15]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

# 汇总行
summary_row = len(items) + 2
ws.merge_cells(f"A{summary_row}:F{summary_row}")
ws.cell(row=summary_row, column=1, value="合计").font = Font(name="微软雅黑", size=11, bold=True)
ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
total_all = sum(qty * price for _, _, _, qty, price in items) / 10000
ws.cell(row=summary_row, column=7, value="总造价").font = Font(name="微软雅黑", size=11, bold=True)
ws.cell(row=summary_row, column=8, value=f"{total_all:.2f}万元").font = Font(name="微软雅黑", size=11, bold=True, color="C62828")

output = r"c:\Users\Administrator\Documents\GitHub\cyq2026\测试清单_破碎楼项目.xlsx"
wb.save(output)
print(f"✅ 测试清单已生成：{output}")
print(f"   共 {len(items)} 条子目，总造价 {total_all:.2f} 万元")
print(f"   故意设置了 8 条偏差项（6条严重🔴 + 1条注意🟡），用于验证审核引擎")
