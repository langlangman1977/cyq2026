from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sqlite3
import csv
import io
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = os.urandom(24).hex()
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)

CORS(app)

DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'zaojiatong.db')


# ==================== 数据库 ====================

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    conn = get_db()
    cur = conn.cursor()

    # 用户表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            company TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            role TEXT DEFAULT 'user',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        )
    """)

    # 项目表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            region TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)

    # 材料查询记录表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS queries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            project_id INTEGER,
            material_name TEXT NOT NULL,
            spec TEXT DEFAULT '',
            unit TEXT DEFAULT '',
            region TEXT DEFAULT '',
            price TEXT DEFAULT '',
            price_unit TEXT DEFAULT '',
            source TEXT DEFAULT '',
            note TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    """)

    # 检查有没有默认管理员，没有则创建
    admin = cur.execute("SELECT id FROM users WHERE username='admin'").fetchone()
    if not admin:
        cur.execute(
            "INSERT INTO users (username, password_hash, company, role) VALUES (?, ?, ?, ?)",
            ('admin', generate_password_hash('admin123'), '中联永信工程管理', 'admin')
        )

    conn.commit()
    conn.close()


# ==================== （无登录验证） ====================


# ==================== 页面路由 ====================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/login")
def login_page():
    return render_template("index.html")


@app.route("/register")
def register_page():
    return render_template("index.html")


# ==================== API：用户 ====================

@app.route("/api/register", methods=["POST"])
def api_register():
    data = request.get_json()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    company = (data.get("company") or "").strip()

    if not username or not password:
        return jsonify({"error": "用户名和密码不能为空"}), 400
    if len(username) < 2:
        return jsonify({"error": "用户名至少2个字符"}), 400
    if len(password) < 6:
        return jsonify({"error": "密码至少6位"}), 400

    db = get_db()
    exist = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if exist:
        db.close()
        return jsonify({"error": "用户名已存在"}), 400

    db.execute(
        "INSERT INTO users (username, password_hash, company) VALUES (?, ?, ?)",
        (username, generate_password_hash(password), company)
    )
    db.commit()
    db.close()

    return jsonify({"message": "注册成功，请登录"})


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()

    if not username or not password:
        return jsonify({"error": "请输入用户名和密码"}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    db.close()

    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "用户名或密码错误"}), 401

    session['user_id'] = user["id"]
    session['username'] = user["username"]
    session['company'] = user["company"]
    session['role'] = user["role"]
    session.permanent = True

    return jsonify({
        "message": "登录成功",
        "user": {
            "id": user["id"],
            "username": user["username"],
            "company": user["company"],
            "role": user["role"]
        }
    })


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"message": "已退出"})


@app.route("/api/me", methods=["GET"])
def api_me():
    db = get_db()
    user = db.execute("SELECT id, username, company, role, created_at FROM users WHERE id=?", (1,)).fetchone()
    db.close()
    if not user:
        return jsonify({"error": "用户不存在"}), 404
    return jsonify(dict(user))


# ==================== API：项目 ====================

@app.route("/api/projects", methods=["GET"])
def api_projects():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM projects WHERE user_id=? ORDER BY created_at DESC",
        (1,)
    ).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/projects", methods=["POST"])
def api_create_project():
    data = request.get_json()
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "项目名称不能为空"}), 400

    db = get_db()
    cur = db.execute(
        "INSERT INTO projects (user_id, name, description, region) VALUES (?, ?, ?, ?)",
        (1, name, data.get("description", ""), data.get("region", ""))
    )
    db.commit()
    pid = cur.lastrowid
    db.close()

    return jsonify({"id": pid, "message": "项目创建成功"})


@app.route("/api/projects/<int:pid>", methods=["DELETE"])
def api_delete_project(pid):
    db = get_db()
    db.execute("DELETE FROM projects WHERE id=? AND user_id=?", (pid, 1))
    db.commit()
    db.close()
    return jsonify({"message": "已删除"})


# ==================== API：材料查询 ====================

@app.route("/api/search", methods=["POST"])
def search():
    data = request.get_json()
    keyword = (data.get("keyword") or "").strip()
    region = (data.get("region") or "").strip()
    project_id = data.get("project_id")

    if not keyword:
        return jsonify({"error": "请输入材料名称"}), 400

    results = search_price(keyword, region)
    save_query(keyword, "", "", region, results, project_id)

    return jsonify({"keyword": keyword, "region": region, "results": results})


@app.route("/api/batch", methods=["POST"])
def batch():
    if "file" not in request.files:
        return jsonify({"error": "请上传文件"}), 400

    file = request.files["file"]
    filename = file.filename.lower()

    try:
        if filename.endswith(".csv"):
            rows = parse_csv(file)
        elif filename.endswith((".xlsx", ".xls")):
            rows = parse_excel(file)
        elif filename.endswith(".pdf"):
            rows = parse_pdf(file)
        else:
            return jsonify({"error": "仅支持 .csv / .xlsx / .xls / .pdf 格式"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {str(e)}"}), 400

    if not rows:
        return jsonify({"error": "文件中没有数据"}), 400

    region = request.form.get("region", "").strip()
    project_id = request.form.get("project_id")

    all_results = []
    for i, row in enumerate(rows):
        name = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        spec = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        unit = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if not name:
            continue
        kw = f"{name} {spec}".strip()
        results = search_price(kw, region)
        save_query(name, spec, unit, region, results, project_id)
        all_results.append({
            "seq": i + 1,
            "name": name,
            "spec": spec,
            "unit": unit,
            "keyword": kw,
            "results": results
        })

    return jsonify({"total": len(all_results), "items": all_results})


@app.route("/api/export", methods=["POST"])
def export():
    data = request.get_json()
    items = data.get("items", [])
    project_id = data.get("project_id")

    # 如果有 project_id，从数据库读取该项目所有查询
    if project_id:
        db = get_db()
        rows = db.execute(
            "SELECT * FROM queries WHERE user_id=? AND project_id=? ORDER BY created_at",
            (1, project_id)
        ).fetchall()
        db.close()
        items = export_from_db(rows)
    elif not items:
        return jsonify({"error": "没有数据可导出"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "材料价格查询结果"
    ws.append(["序号", "材料名称", "规格型号", "单位", "查询地区",
               "参考价格", "价格单位", "信息来源", "备注", "查询时间"])

    for i, item in enumerate(items):
        results = item.get("results", [])
        if results:
            for r in results:
                ws.append([
                    i + 1,
                    item.get("name", ""),
                    item.get("spec", ""),
                    item.get("unit", ""),
                    item.get("region", ""),
                    r.get("price", ""),
                    r.get("unit", ""),
                    r.get("source", ""),
                    r.get("note", ""),
                    datetime.now().strftime("%Y-%m-%d %H:%M")
                ])
        else:
            ws.append([
                i + 1,
                item.get("name", ""),
                item.get("spec", ""),
                item.get("unit", ""),
                item.get("region", ""),
                "未查到", "", "", "",
                datetime.now().strftime("%Y-%m-%d %H:%M")
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"材料价格查询_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


@app.route("/api/history", methods=["GET"])
def history():
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    project_id = request.args.get("project_id")

    db = get_db()
    where = "WHERE user_id=?"
    params = [1]
    if project_id:
        where += " AND project_id=?"
        params.append(project_id)

    total = db.execute(f"SELECT COUNT(*) FROM queries {where}", params).fetchone()[0]
    rows = db.execute(
        f"SELECT * FROM queries {where} ORDER BY created_at DESC LIMIT ? OFFSET ?",
        params + [per_page, (page - 1) * per_page]
    ).fetchall()
    db.close()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "items": [dict(r) for r in rows]
    })


# ==================== API：清单审核 ====================

import tempfile, shutil

@app.route("/api/audit", methods=["POST"])
def api_audit():
    """接收上传的清单文件，运行审核引擎，返回审核结果"""
    if "file" not in request.files:
        return jsonify({"error": "请上传清单文件"}), 400

    file = request.files["file"]
    building_type = request.form.get("building_type", "工业建筑").strip()
    area_str = request.form.get("area", "0").strip()
    project_name = request.form.get("project_name", "").strip()

    area = 0.0
    try:
        area = float(area_str)
    except ValueError:
        pass

    # 保存上传文件到临时目录
    tmp_dir = tempfile.mkdtemp()
    filename = file.filename or "uploaded_list.xlsx"
    safe_name = "".join(c for c in filename if c.isalnum() or c in "._-")
    filepath = os.path.join(tmp_dir, safe_name)
    file.save(filepath)

    try:
        # PDF 文件先转 Excel 再审核
        is_pdf = filename.lower().endswith(".pdf")
        if is_pdf:
            rows = parse_pdf(filepath)
            if not rows:
                return jsonify({"error": "PDF 文件中未检测到表格数据，请确保清单以表格形式存在"}), 400
            # 将提取的行写入临时 Excel 供 Auditor 解析
            pdf_excel = os.path.join(tmp_dir, safe_name.replace(".pdf", "_extracted.xlsx"))
            wb_out = Workbook()
            ws_out = wb_out.active
            for r in rows:
                ws_out.append(r)
            wb_out.save(pdf_excel)
            filepath = pdf_excel

        from zaojia_auditor import Auditor
        auditor = Auditor(filepath)
        if not auditor.parse():
            return jsonify({"error": "无法解析清单文件，请检查格式" + ("（PDF 已提取表格但可能表头不规范）" if is_pdf else "")}), 400

        auditor.audit_prices()
        auditor.analyze_indicators(building_type, area)

        # 保存审核报告
        report_name = f"审核报告_{safe_name}"
        report_path = os.path.join(tmp_dir, report_name)
        if not report_path.endswith(".xlsx"):
            report_path += ".xlsx"
        auditor.save_report(report_path)

        # 如果指定了项目名，也存入记忆库
        if project_name:
            import list_memory
            list_memory.remember_list(filepath, project_name, building_type, area)

        # 构建返回值
        price_findings = [f for f in auditor.findings if f.get("type") == "单价审核"]
        high_count = sum(1 for f in price_findings if f.get("level") == "high")
        med_count = sum(1 for f in price_findings if f.get("level") == "medium")

        # 复制报告到 static 目录供下载
        static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
        os.makedirs(static_dir, exist_ok=True)
        static_report = os.path.join(static_dir, report_name)
        shutil.copy2(report_path, static_report)

        return jsonify({
            "item_count": len(price_findings),
            "total_price": auditor.total_price,
            "high_count": high_count,
            "med_count": med_count,
            "indicators": {k: {"value": v.get("value", v) if isinstance(v, dict) else v,
                               "range": v.get("range", "") if isinstance(v, dict) else "",
                               "ok": v.get("ok", True)} for k, v in auditor.indicators.items()},
            "findings": [{
                "name": f.get("name", ""),
                "unit": f.get("unit", ""),
                "qty": f.get("qty", 0),
                "price": f.get("price", 0),
                "ref_range": f.get("ref_range", ""),
                "deviation": f.get("deviation", 0),
                "level": f.get("level", "normal"),
            } for f in price_findings if f.get("level") != "normal"],
            "report_file": report_name,
            "report_url": f"/static/{report_name}",
        })

    except Exception as e:
        return jsonify({"error": f"审核失败: {str(e)}"}), 500
    finally:
        # 清理临时目录（保留报告）
        try:
            for f in os.listdir(tmp_dir):
                fp = os.path.join(tmp_dir, f)
                if os.path.isfile(fp) and not f.startswith("审核报告"):
                    os.remove(fp)
        except:
            pass


@app.route("/api/download_report", methods=["GET"])
def api_download_report():
    """下载审核报告"""
    filename = request.args.get("file", "").strip()
    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
    path = os.path.join(static_dir, filename)
    if not os.path.exists(path):
        return jsonify({"error": "报告文件不存在"}), 404
    return send_file(path, as_attachment=True, download_name=filename)


@app.route("/api/memory_stats", methods=["GET"])
def api_memory_stats():
    """获取记忆库统计"""
    import list_memory
    stats = list_memory.get_memory_stats()
    return jsonify(stats)


@app.route("/api/memory_projects", methods=["GET"])
def api_memory_projects():
    """获取已审项目列表"""
    import list_memory
    projects = list_memory.get_all_projects()
    return jsonify({"projects": projects})


@app.route("/api/memory_recall", methods=["GET"])
def api_memory_recall():
    """召回相似子目价格记忆"""
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"items": []})
    import list_memory
    items = list_memory.recall_similar(name)
    return jsonify({"items": items})

def save_query(name, spec, unit, region, results, project_id=None):
    """保存查询记录到数据库"""
    db = get_db()
    pid = project_id if project_id else None
    if results:
        for r in results:
            db.execute(
                """INSERT INTO queries (user_id, project_id, material_name, spec, unit, region,
                   price, price_unit, source, note) VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (1, pid, name, spec, unit, region,
                 r.get("price", ""), r.get("unit", ""),
                 r.get("source", ""), r.get("note", ""))
            )
    else:
        db.execute(
            """INSERT INTO queries (user_id, project_id, material_name, spec, unit, region)
               VALUES (?,?,?,?,?,?)""",
            (1, pid, name, spec, unit, region)
        )
    db.commit()
    db.close()


def search_price(keyword, region=""):
    """搜索材料价格 - 三层查找策略

    1. 本地价格库（覆盖60+种主要建材）
    2. 地区价格调整（按地区系数修正）
    3. 找不到返回空，由调用方决定是否联网搜
    """
    results = []
    kw_lower = keyword.lower().replace(" ", "")

    # 关键字→价格匹配（覆盖主要建材品类）
    sample = {
        # === 混凝土（泵送商品砼）===
        "c15": [{"price_lo": "365", "price_hi": "395", "unit": "元/m³", "source": "造价通"}],
        "c20": [{"price_lo": "390", "price_hi": "420", "unit": "元/m³", "source": "造价通"}],
        "c25": [{"price_lo": "415", "price_hi": "445", "unit": "元/m³", "source": "造价通"}],
        "c30": [{"price_lo": "440", "price_hi": "500", "unit": "元/m³", "source": "造价通"}],
        "c35": [{"price_lo": "470", "price_hi": "530", "unit": "元/m³", "source": "造价通"}],
        "c40": [{"price_lo": "500", "price_hi": "560", "unit": "元/m³", "source": "造价通"}],
        "c45": [{"price_lo": "540", "price_hi": "600", "unit": "元/m³", "source": "造价通"}],
        "c50": [{"price_lo": "580", "price_hi": "650", "unit": "元/m³", "source": "造价通"}],
        # === 钢筋 ===
        "hpb300": [{"price_lo": "3700", "price_hi": "4100", "unit": "元/吨", "source": "Mysteel"}],
        "hrb400": [{"price_lo": "3580", "price_hi": "4050", "unit": "元/吨", "source": "Mysteel"}],
        "hrb400e": [{"price_lo": "3650", "price_hi": "4120", "unit": "元/吨", "source": "Mysteel"}],
        "hrb500": [{"price_lo": "3900", "price_hi": "4350", "unit": "元/吨", "source": "Mysteel"}],
        "冷轧带肋": [{"price_lo": "3850", "price_hi": "4250", "unit": "元/吨", "source": "Mysteel"}],
        "钢绞线": [{"price_lo": "5200", "price_hi": "5800", "unit": "元/吨", "source": "Mysteel"}],
        # === 钢结构 ===
        "q235": [{"price_lo": "4200", "price_hi": "4800", "unit": "元/吨", "source": "Mysteel"}],
        "q355": [{"price_lo": "4500", "price_hi": "5200", "unit": "元/吨", "source": "Mysteel"}],
        "q390": [{"price_lo": "4700", "price_hi": "5400", "unit": "元/吨", "source": "Mysteel"}],
        "q420": [{"price_lo": "4900", "price_hi": "5600", "unit": "元/吨", "source": "Mysteel"}],
        "工字钢": [{"price_lo": "4300", "price_hi": "4900", "unit": "元/吨", "source": "Mysteel"}],
        "h型钢": [{"price_lo": "4500", "price_hi": "5100", "unit": "元/吨", "source": "Mysteel"}],
        "槽钢": [{"price_lo": "4250", "price_hi": "4850", "unit": "元/吨", "source": "Mysteel"}],
        "角钢": [{"price_lo": "4200", "price_hi": "4750", "unit": "元/吨", "source": "Mysteel"}],
        "钢板": [{"price_lo": "4300", "price_hi": "4900", "unit": "元/吨", "source": "Mysteel"}],
        # === 水泥/粉煤灰/外加剂 ===
        "p.o42.5": [{"price_lo": "480", "price_hi": "550", "unit": "元/吨", "source": "造价通"}],
        "p.o52.5": [{"price_lo": "550", "price_hi": "650", "unit": "元/吨", "source": "造价通"}],
        "普通硅酸盐": [{"price_lo": "480", "price_hi": "560", "unit": "元/吨", "source": "造价通"}],
        "矿渣硅酸盐": [{"price_lo": "420", "price_hi": "500", "unit": "元/吨", "source": "造价通"}],
        "粉煤灰": [{"price_lo": "120", "price_hi": "180", "unit": "元/吨", "source": "造价通"}],
        "减水剂": [{"price_lo": "6500", "price_hi": "7500", "unit": "元/吨", "source": "造价通"}],
        "防水剂": [{"price_lo": "8500", "price_hi": "11000", "unit": "元/吨", "source": "造价通"}],
        # === 砂/石/骨料 ===
        "中砂": [{"price_lo": "80", "price_hi": "120", "unit": "元/m³", "source": "造价通"}],
        "细砂": [{"price_lo": "70", "price_hi": "110", "unit": "元/m³", "source": "造价通"}],
        "粗砂": [{"price_lo": "85", "price_hi": "125", "unit": "元/m³", "source": "造价通"}],
        "机制砂": [{"price_lo": "75", "price_hi": "115", "unit": "元/m³", "source": "造价通"}],
        "碎石5-10": [{"price_lo": "90", "price_hi": "130", "unit": "元/m³", "source": "造价通"}],
        "碎石5-20": [{"price_lo": "95", "price_hi": "135", "unit": "元/m³", "source": "造价通"}],
        "碎石5-31.5": [{"price_lo": "90", "price_hi": "130", "unit": "元/m³", "source": "造价通"}],
        "玄武岩": [{"price_lo": "180", "price_hi": "240", "unit": "元/m³", "source": "造价通"}],
        "页岩陶粒": [{"price_lo": "320", "price_hi": "380", "unit": "元/m³", "source": "造价通"}],
        # === 砌体 ===
        "加气块": [{"price_lo": "280", "price_hi": "350", "unit": "元/m³", "source": "造价通"}],
        "加气混凝土砌块": [{"price_lo": "280", "price_hi": "350", "unit": "元/m³", "source": "造价通"}],
        "页岩砖": [{"price_lo": "380", "price_hi": "450", "unit": "元/m³", "source": "造价通"}],
        "页岩多孔砖": [{"price_lo": "350", "price_hi": "420", "unit": "元/m³", "source": "造价通"}],
        "煤矸石砖": [{"price_lo": "320", "price_hi": "380", "unit": "元/m³", "source": "造价通"}],
        "烧结砖": [{"price_lo": "380", "price_hi": "450", "unit": "元/千块", "source": "造价通"}],
        # === 防水材料 ===
        "sbs": [{"price_lo": "24", "price_hi": "35", "unit": "元/m²", "source": "造价通"}],
        "防水卷材": [{"price_lo": "24", "price_hi": "35", "unit": "元/m²", "source": "造价通"}],
        "自粘卷材": [{"price_lo": "32", "price_hi": "45", "unit": "元/m²", "source": "造价通"}],
        "js防水": [{"price_lo": "75", "price_hi": "95", "unit": "元/m²", "source": "造价通"}],
        "聚氨酯": [{"price_lo": "14", "price_hi": "18", "unit": "元/kg", "source": "造价通"}],
        "js聚合物": [{"price_lo": "16", "price_hi": "22", "unit": "元/kg", "source": "造价通"}],
        "防水涂料": [{"price_lo": "14", "price_hi": "22", "unit": "元/kg", "source": "造价通"}],
        # === 保温材料 ===
        "xps": [{"price_lo": "50", "price_hi": "65", "unit": "元/m²", "source": "造价通"}],
        "eps": [{"price_lo": "38", "price_hi": "48", "unit": "元/m²", "source": "造价通"}],
        "岩棉": [{"price_lo": "50", "price_hi": "60", "unit": "元/m²", "source": "造价通"}],
        "玻璃棉": [{"price_lo": "45", "price_hi": "55", "unit": "元/m²", "source": "造价通"}],
        "挤塑板": [{"price_lo": "50", "price_hi": "65", "unit": "元/m²", "source": "造价通"}],
        "保温板": [{"price_lo": "48", "price_hi": "62", "unit": "元/m²", "source": "造价通"}],
        # === 涂料油漆 ===
        "乳胶漆": [{"price_lo": "12", "price_hi": "18", "unit": "元/kg", "source": "造价通"}],
        "内墙漆": [{"price_lo": "12", "price_hi": "18", "unit": "元/kg", "source": "造价通"}],
        "外墙漆": [{"price_lo": "20", "price_hi": "32", "unit": "元/kg", "source": "造价通"}],
        "真石漆": [{"price_lo": "55", "price_hi": "85", "unit": "元/m²", "source": "造价通"}],
        "防火涂料": [{"price_lo": "12", "price_hi": "18", "unit": "元/kg", "source": "造价通"}],
        "防锈漆": [{"price_lo": "16", "price_hi": "24", "unit": "元/kg", "source": "造价通"}],
        "环氧漆": [{"price_lo": "32", "price_hi": "48", "unit": "元/kg", "source": "造价通"}],
        # === 门窗 ===
        "铝合金门": [{"price_lo": "520", "price_hi": "680", "unit": "元/m²", "source": "造价通"}],
        "铝合金窗": [{"price_lo": "480", "price_hi": "680", "unit": "元/m²", "source": "造价通"}],
        "断桥铝": [{"price_lo": "580", "price_hi": "780", "unit": "元/m²", "source": "造价通"}],
        "塑钢窗": [{"price_lo": "350", "price_hi": "450", "unit": "元/m²", "source": "造价通"}],
        "防火门": [{"price_lo": "450", "price_hi": "650", "unit": "元/m²", "source": "造价通"}],
        "防盗门": [{"price_lo": "450", "price_hi": "650", "unit": "元/m²", "source": "造价通"}],
        "钢质门": [{"price_lo": "400", "price_hi": "600", "unit": "元/m²", "source": "造价通"}],
        "木门": [{"price_lo": "1200", "price_hi": "2500", "unit": "元/樘", "source": "造价通"}],
        # === 瓷砖/石材 ===
        "瓷砖": [{"price_lo": "45", "price_hi": "120", "unit": "元/m²", "source": "造价通"}],
        "地砖": [{"price_lo": "45", "price_hi": "120", "unit": "元/m²", "source": "造价通"}],
        "墙砖": [{"price_lo": "35", "price_hi": "80", "unit": "元/m²", "source": "造价通"}],
        "抛光砖": [{"price_lo": "65", "price_hi": "150", "unit": "元/m²", "source": "造价通"}],
        "釉面砖": [{"price_lo": "45", "price_hi": "100", "unit": "元/m²", "source": "造价通"}],
        "石材": [{"price_lo": "280", "price_hi": "1200", "unit": "元/m²", "source": "造价通"}],
        "花岗岩": [{"price_lo": "180", "price_hi": "600", "unit": "元/m²", "source": "造价通"}],
        "大理石": [{"price_lo": "350", "price_hi": "1500", "unit": "元/m²", "source": "造价通"}],
        # === 电线电缆 ===
        "bv1.5": [{"price_lo": "1.8", "price_hi": "2.2", "unit": "元/m", "source": "造价通"}],
        "bv2.5": [{"price_lo": "3.8", "price_hi": "4.8", "unit": "元/m", "source": "造价通"}],
        "bv4": [{"price_lo": "6.5", "price_hi": "8.0", "unit": "元/m", "source": "造价通"}],
        "bv6": [{"price_lo": "9.5", "price_hi": "11", "unit": "元/m", "source": "造价通"}],
        "bv10": [{"price_lo": "16", "price_hi": "20", "unit": "元/m", "source": "造价通"}],
        "yjv": [{"price_lo": "80", "price_hi": "120", "unit": "元/m", "source": "造价通"}],
        "yjv22": [{"price_lo": "95", "price_hi": "140", "unit": "元/m", "source": "造价通"}],
        "kvv": [{"price_lo": "12", "price_hi": "22", "unit": "元/m", "source": "造价通"}],
        "电力电缆": [{"price_lo": "80", "price_hi": "150", "unit": "元/m", "source": "造价通"}],
        "控制电缆": [{"price_lo": "12", "price_hi": "25", "unit": "元/m", "source": "造价通"}],
        # === 管材 ===
        "ppr": [{"price_lo": "6", "price_hi": "10", "unit": "元/m", "source": "造价通"}],
        "ppr20": [{"price_lo": "6", "price_hi": "10", "unit": "元/m", "source": "造价通"}],
        "ppr25": [{"price_lo": "10", "price_hi": "16", "unit": "元/m", "source": "造价通"}],
        "pvc-u": [{"price_lo": "15", "price_hi": "25", "unit": "元/m", "source": "造价通"}],
        "pvc110": [{"price_lo": "15", "price_hi": "25", "unit": "元/m", "source": "造价通"}],
        "pvc160": [{"price_lo": "28", "price_hi": "40", "unit": "元/m", "source": "造价通"}],
        "镀锌钢管": [{"price_lo": "45", "price_hi": "65", "unit": "元/m", "source": "造价通"}],
        "焊接钢管": [{"price_lo": "38", "price_hi": "55", "unit": "元/m", "source": "造价通"}],
        "无缝钢管": [{"price_lo": "5500", "price_hi": "6800", "unit": "元/吨", "source": "造价通"}],
        "pe管": [{"price_lo": "12", "price_hi": "28", "unit": "元/m", "source": "造价通"}],
        "ppr管": [{"price_lo": "6", "price_hi": "10", "unit": "元/m", "source": "造价通"}],
        "pvc管": [{"price_lo": "15", "price_hi": "25", "unit": "元/m", "source": "造价通"}],
        # === 玻璃 ===
        "中空玻璃": [{"price_lo": "80", "price_hi": "120", "unit": "元/m²", "source": "造价通"}],
        "钢化玻璃": [{"price_lo": "120", "price_hi": "180", "unit": "元/m²", "source": "造价通"}],
        "夹胶玻璃": [{"price_lo": "150", "price_hi": "220", "unit": "元/m²", "source": "造价通"}],
        "low-e玻璃": [{"price_lo": "180", "price_hi": "260", "unit": "元/m²", "source": "造价通"}],
        "镀膜玻璃": [{"price_lo": "160", "price_hi": "230", "unit": "元/m²", "source": "造价通"}],
        # === 装饰板材 ===
        "石膏板": [{"price_lo": "15", "price_hi": "22", "unit": "元/m²", "source": "造价通"}],
        "矿棉板": [{"price_lo": "28", "price_hi": "45", "unit": "元/m²", "source": "造价通"}],
        "硅钙板": [{"price_lo": "22", "price_hi": "38", "unit": "元/m²", "source": "造价通"}],
        "轻钢龙骨": [{"price_lo": "10", "price_hi": "18", "unit": "元/m²", "source": "造价通"}],
        "木龙骨": [{"price_lo": "1200", "price_hi": "1800", "unit": "元/m³", "source": "造价通"}],
        "细木工板": [{"price_lo": "35", "price_hi": "55", "unit": "元/m²", "source": "造价通"}],
        "胶合板": [{"price_lo": "12", "price_hi": "20", "unit": "元/m²", "source": "造价通"}],
        # === 市政材料 ===
        "球墨铸铁管": [{"price_lo": "5800", "price_hi": "6800", "unit": "元/吨", "source": "造价通"}],
        "钢筋混凝土管": [{"price_lo": "380", "price_hi": "520", "unit": "元/m", "source": "造价通"}],
        "hdpe双壁波纹管": [{"price_lo": "38", "price_hi": "85", "unit": "元/m", "source": "造价通"}],
        "检查井": [{"price_lo": "850", "price_hi": "1500", "unit": "元/座", "source": "造价通"}],
    }

    # 地区价格系数（基于 2026 年区域差异）
    region_coefficients = {
        "北京": 1.15, "上海": 1.18, "广州": 1.20, "深圳": 1.22,
        "杭州": 1.15, "南京": 1.10, "苏州": 1.12, "成都": 1.05,
        "武汉": 1.05, "西安": 1.02, "重庆": 1.03, "天津": 1.08,
        "青岛": 1.06, "济南": 1.03, "郑州": 1.00, "洛阳": 0.95,
        "南宁": 1.00, "桂林": 0.92, "北海": 1.00, "海口": 1.05,
        "新疆": 1.08, "西藏": 1.18, "东北": 1.02,
        # 河南各地
        "河南": 1.00, "郑州": 1.00, "开封": 0.95, "洛阳": 0.95,
        "平顶山": 0.93, "安阳": 0.94, "鹤壁": 0.93, "新乡": 0.95,
        "焦作": 0.96, "濮阳": 0.94, "许昌": 0.94, "漯河": 0.94,
        "三门峡": 0.93, "南阳": 0.93, "商丘": 0.93, "信阳": 0.93,
        "周口": 0.93, "驻马店": 0.93, "济源": 0.95,
        # 广西
        "广西": 1.00, "南宁": 1.00, "柳州": 0.96, "桂林": 0.92,
        "梧州": 0.94, "北海": 1.00, "钦州": 0.98, "防城港": 0.98,
    }

    # 模糊匹配
    matched_key = None
    for k, v in sample.items():
        if k in kw_lower:
            results = v
            matched_key = k
            break

    # 应用地区系数
    if results and region:
        coef = region_coefficients.get(region, 1.0)
        # 沿海地区混凝土+15%/钢筋+5% 特殊调整
        is_coastal = region in ["广西", "北海", "防城港", "广州", "深圳", "上海", "杭州", "宁波", "厦门"]
        for r in results:
            lo = float(r.get("price_lo", 0))
            hi = float(r.get("price_hi", 0))
            unit = r.get("unit", "")
            if is_coastal:
                if "m³" in unit:  # 混凝土类
                    coef_eff = max(coef, 1.10)
                elif "吨" in unit:  # 钢筋/钢材
                    coef_eff = max(coef, 1.05)
                else:
                    coef_eff = coef
            else:
                coef_eff = coef

            r["price_lo"] = round(lo * coef_eff, 1) if coef_eff != 1.0 else lo
            r["price_hi"] = round(hi * coef_eff, 1) if coef_eff != 1.0 else hi
            r["region"] = region
            r["note"] = f"河南基准·{matched_key}·地区系数{coef_eff}·2026年7月参考"
            r["price"] = f"{r['price_lo']}-{r['price_hi']}"

    return results


def export_from_db(rows):
    """将数据库记录转为导出格式"""
    groups = {}
    for r in rows:
        key = (r["material_name"], r["spec"], r["unit"])
        if key not in groups:
            groups[key] = {
                "name": r["material_name"],
                "spec": r["spec"],
                "unit": r["unit"],
                "region": r["region"],
                "results": []
            }
        if r["price"]:
            groups[key]["results"].append({
                "price": r["price"],
                "unit": r["price_unit"],
                "source": r["source"],
                "note": r["note"]
            })
    return list(groups.values())


# ==================== 文件解析 ====================

def parse_csv(file):
    content = file.read().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    rows = []
    for i, row in enumerate(reader):
        if i == 0:
            continue
        if any(cell.strip() for cell in row if cell):
            rows.append(row)
    return rows


def parse_excel(file):
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if any(str(c).strip() for c in row if c):
            rows.append(row)
    return rows


def parse_pdf(file):
    """从 PDF 中提取表格数据"""
    import pdfplumber
    rows = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for r in table:
                    if r and any(str(c).strip() for c in r if c):
                        rows.append([str(c) if c else "" for c in r])
    return rows


# ==================== 启动 ====================

if __name__ == "__main__":
    init_db()
    print("Zaojiatong system started")
    print("    URL: http://127.0.0.1:5000")
    print("    Default admin: admin / admin123")
    app.run(host="0.0.0.0", port=5000, debug=True)
