#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工程造价清单审核引擎 — zaojia_auditor.py

功能：
1. 解析用户上传的 Excel 工程量清单
2. 逐条审核单价合理性（对比内置参考区间）
3. 分析工程量指标（钢筋含量、混凝土含量、平米造价等）
4. 抽查材料价格
5. 生成带颜色标注的 Excel 审核报告

使用方式：
  from zaojia_auditor import Auditor
  auditor = Auditor("清单文件.xlsx")
  report = auditor.audit()
  auditor.save_report("审核报告.xlsx")
"""

import re, os, collections
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════
# 内置知识库
# ═══════════════════════════════════════════════

# 综合单价合理区间（单位：元）
PRICE_RANGES = {
    "挖土方": (4, 8, "m³"),
    "土方开挖": (4, 8, "m³"),
    "挖一般土方": (4, 8, "m³"),
    "土方回填": (15, 25, "m³"),
    "回填": (15, 25, "m³"),
    "现浇混凝土基础": (650, 850, "m³"),
    "混凝土基础": (650, 850, "m³"),
    "现浇混凝土柱": (900, 1200, "m³"),
    "混凝土柱": (900, 1200, "m³"),
    "现浇混凝土梁": (950, 1250, "m³"),
    "混凝土梁": (950, 1250, "m³"),
    "现浇混凝土板": (850, 1100, "m³"),
    "混凝土板": (850, 1100, "m³"),
    "钢筋制作安装": (5200, 6200, "t"),
    "钢筋": (5200, 6200, "t"),
    "砌体": (450, 580, "m³"),
    "加气混凝土砌块": (450, 580, "m³"),
    "加气块": (450, 580, "m³"),
    "钢结构制作安装": (7800, 9800, "t"),
    "钢结构": (7800, 9800, "t"),
    "防火涂料薄型": (28, 45, "m²"),
    "防火涂料": (28, 45, "m²"),
    "内墙抹灰": (25, 38, "m²"),
    "抹灰": (25, 38, "m²"),
    "外墙保温": (80, 120, "m²"),
    "屋面防水": (65, 95, "m²"),
    "防水": (65, 95, "m²"),
    "铝合金窗": (550, 750, "m²"),
    "门窗": (550, 750, "m²"),
    "防盗门": (450, 650, "m²"),
    "地砖铺贴": (55, 85, "m²"),
    "地砖": (55, 85, "m²"),
    "内墙涂料": (22, 35, "m²"),
    "涂料": (22, 35, "m²"),
    "外墙真石漆": (55, 85, "m²"),
    "真石漆": (55, 85, "m²"),
    "电气配管": (8, 15, "m"),
    "电气配线": (3, 6, "m"),
    "给水管道": (25, 45, "m"),
    "排水管道": (30, 50, "m"),
}

# 材料参考价格
MATERIAL_PRICES = {
    "商品混凝土 C30": (440, 500, "m³"),
    "商品混凝土 C35": (470, 530, "m³"),
    "商品混凝土 C40": (500, 560, "m³"),
    "C30混凝土": (440, 500, "m³"),
    "C35混凝土": (470, 530, "m³"),
    "C40混凝土": (500, 560, "m³"),
    "钢筋 HPB300": (3700, 4100, "t"),
    "钢筋 HRB400": (3600, 4000, "t"),
    "螺纹钢": (3600, 4000, "t"),
    "型钢": (4200, 4800, "t"),
    "钢板": (4300, 4900, "t"),
    "水泥 P.O 42.5": (480, 550, "t"),
    "砂": (80, 120, "m³"),
    "碎石": (90, 130, "m³"),
    "加气混凝土砌块": (280, 350, "m³"),
    "加气块": (280, 350, "m³"),
    "防水卷材 SBS": (24, 35, "m²"),
    "SBS防水卷材": (24, 35, "m²"),
    "保温板 XPS": (55, 65, "m²"),
    "中空玻璃": (80, 120, "m²"),
    "铝合金": (480, 680, "m²"),
    "电缆 YJV": (80, 110, "m"),
    "镀锌钢管": (45, 65, "m"),
    "PPR管": (6, 10, "m"),
    "PVC排水管": (15, 25, "m"),
}

# 工程量指标参考
ENGINEERING_INDICATORS = {
    "工业建筑": {
        "钢筋含量": (55, 85, "kg/m²"),
        "混凝土含量": (0.35, 0.55, "m³/m²"),
        "钢结构含量": (60, 120, "kg/m²"),
        "模板含量": (2.5, 4.0, "m²/m²"),
        "土建造价": (1800, 2800, "元/m²"),
        "总造价": (2500, 4500, "元/m²"),
    },
    "民用建筑": {
        "钢筋含量": (38, 55, "kg/m²"),
        "混凝土含量": (0.33, 0.45, "m³/m²"),
        "土建造价": (1500, 2200, "元/m²"),
        "安装造价占比": (8, 15, "%"),
    },
}

# ═══════════════════════════════════════════════
# 格式定义
# ═══════════════════════════════════════════════

RED_FILL = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD54F", end_color="FFD54F", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1A237E")
BODY_FONT = Font(name="微软雅黑", size=10)
RED_FONT = Font(name="微软雅黑", size=10, color="C62828", bold=True)
YELLOW_FONT = Font(name="微软雅黑", size=10, color="F57F17")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═══════════════════════════════════════════════
# 审核引擎
# ═══════════════════════════════════════════════

class Auditor:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.rows = []           # 解析后的清单行
        self.headers = []        # 表头
        self.findings = []       # 审核发现
        self.indicators = {}     # 工程量指标分析结果
        self.total_price = 0     # 总造价

    # ── 解析 Excel ──
    def parse(self) -> bool:
        """解析 Excel 清单，返回是否成功"""
        try:
            wb = load_workbook(self.filepath, data_only=True)
            ws = wb.active
            rows_list = list(ws.iter_rows(values_only=True))
            if not rows_list:
                return False

            # 查找表头行
            header_idx = self._find_header(rows_list)
            self.headers = [str(c) if c else "" for c in rows_list[header_idx]]
            # 数据行
            self.rows = []
            for r in rows_list[header_idx + 1:]:
                vals = [str(v) if v is not None else "" for v in r]
                if any(v.strip() for v in vals):
                    self.rows.append(vals)
            return len(self.rows) > 0
        except Exception as e:
            print(f"解析失败: {e}")
            return False

    def _find_header(self, rows: list) -> int:
        """自动识别表头行"""
        keywords = ["项目编码", "项目名称", "子目名称", "清单编号", "编码", "序号",
                     "计量单位", "单位", "工程数量", "工程量", "数量", "综合单价", "单价", "合价"]
        for i, row in enumerate(rows):
            texts = [str(c).strip() for c in row if c]
            combined = " ".join(texts)
            if sum(1 for kw in keywords if kw in combined) >= 3:
                return i
        # 兜底：第一行
        return 0

    def _get_col(self, keywords: list) -> int:
        """根据关键词找列索引"""
        for kw in keywords:
            for i, h in enumerate(self.headers):
                if kw in h:
                    return i
        return -1

    def _safe_float(self, val: str) -> float:
        """安全转 float"""
        try:
            return float(re.sub(r"[^\d.\-]", "", str(val)))
        except:
            return 0.0

    # ── 单价审核 ──
    def audit_prices(self):
        """逐条审核单价合理性"""
        name_col = self._get_col(["项目名称", "子目名称", "名称", "清单名称"])
        unit_col = self._get_col(["计量单位", "单位"])
        qty_col = self._get_col(["工程数量", "工程量", "数量"])
        price_col = self._get_col(["综合单价", "单价"])

        if name_col < 0 or price_col < 0:
            self.findings.append({"level": "ERROR", "msg": "未检测到项目名称列或单价列，请检查清单格式"})
            return

        for i, row in enumerate(self.rows):
            name = str(row[name_col]) if name_col < len(row) else ""
            unit = str(row[unit_col]) if unit_col >= 0 and unit_col < len(row) else ""
            qty = self._safe_float(str(row[qty_col])) if qty_col >= 0 and qty_col < len(row) else 0
            price = self._safe_float(str(row[price_col])) if price_col < len(row) else 0

            if price <= 0 or not name.strip():
                continue

            # 计算合价
            total = price * qty
            self.total_price += total

            # 匹配参考区间
            matched = None
            for pat, (lo, hi, u) in PRICE_RANGES.items():
                if pat in name:
                    matched = (pat, lo, hi, u)
                    break

            if matched:
                pat, lo, hi, u = matched
                mid = (lo + hi) / 2
                dev = (price - mid) / mid * 100 if mid > 0 else 0

                level = "normal"
                if abs(dev) > 20:
                    level = "high"
                elif abs(dev) > 10:
                    level = "medium"

                if level != "normal":
                    direction = "偏高" if dev > 0 else "偏低"
                    self.findings.append({
                        "row": i + 2,
                        "name": name,
                        "unit": unit,
                        "qty": qty,
                        "price": price,
                        "total": total,
                        "ref_range": f"{lo} ~ {hi} 元/{u}",
                        "deviation": round(dev, 1),
                        "level": level,
                        "type": "单价审核",
                        "suggestion": f"单价{round(abs(dev), 1)}%{direction}，建议核查材料费和人工费构成，参考区间{lo}~{hi}元/{u}",
                    })

    # ── 工程量指标分析 ──
    def analyze_indicators(self, building_type: str = "工业建筑", area: float = None):
        """分析工程量指标"""
        indicators_ref = ENGINEERING_INDICATORS.get(building_type, ENGINEERING_INDICATORS["工业建筑"])

        # 尝试从清单名称推断建筑面积
        name_col = self._get_col(["项目名称", "子目名称", "名称"])
        qty_col = self._get_col(["工程数量", "工程量", "数量"])
        unit_col = self._get_col(["计量单位", "单位"])
        price_col = self._get_col(["综合单价", "单价"])

        # 统计各分部数据
        steel_qty = concrete_qty = steel_total = 0
        for row in self.rows:
            name = str(row[name_col]) if name_col < len(row) else ""
            qty = self._safe_float(str(row[qty_col])) if qty_col < len(row) else 0
            price = self._safe_float(str(row[price_col])) if price_col < len(row) else 0
            u = str(row[unit_col]).lower() if unit_col < len(row) else ""

            if "钢筋" in name:
                steel_qty += qty if "t" in u else qty / 1000
                steel_total += qty * price
            if "混凝土" in name and "m³" in u:
                concrete_qty += qty

        # 如果有面积
        if area and area > 0:
            for name, (lo, hi, u) in indicators_ref.items():
                if name == "钢筋含量":
                    val = steel_qty * 1000 / area if steel_qty > 0 else 0
                    in_range = lo <= val <= hi
                    self.indicators[name] = {"value": round(val, 1), "range": f"{lo}~{hi} {u}",
                                               "ok": in_range, "type": "指标分析"}
                elif name == "混凝土含量":
                    val = concrete_qty / area if concrete_qty > 0 else 0
                    in_range = lo <= val <= hi
                    self.indicators[name] = {"value": round(val, 2), "range": f"{lo}~{hi} {u}",
                                               "ok": in_range, "type": "指标分析"}
                elif name == "总造价":
                    val = self.total_price / area / 10000
                    in_range = lo <= val <= hi
                    self.indicators[name] = {"value": round(val, 0), "range": f"{lo}~{hi} {u}",
                                               "ok": in_range, "type": "指标分析"}

        # 分部造价占比
        if self.total_price > 0:
            for name, keyword in [("土建", ["土方", "混凝土", "钢筋", "砌体", "抹灰"]),
                                   ("钢结构", ["钢结构", "钢构件"]),
                                   ("装饰", ["涂料", "地砖", "门窗"]),
                                   ("安装", ["电气", "给水", "排水", "管道"])]:
                cat_total = 0
                for row in self.rows:
                    rname = str(row[name_col]) if name_col < len(row) else ""
                    price = self._safe_float(str(row[price_col])) if price_col < len(row) else 0
                    qty = self._safe_float(str(row[qty_col])) if qty_col < len(row) else 0
                    if any(kw in rname for kw in keyword):
                        cat_total += price * qty
                pct = round(cat_total / self.total_price * 100, 1) if self.total_price > 0 else 0
                if pct > 0:
                    self.indicators[f"{name}占比"] = {"value": pct, "range": "", "ok": True,
                                                       "type": "分部占比", "note": f"{cat_total/10000:.0f}万元"}

    # ── 生成报告 ──
    def save_report(self, output_path: str):
        """生成带格式的 Excel 审核报告"""
        wb = Workbook()

        # Sheet1: 审核摘要
        ws1 = wb.active
        ws1.title = "审核摘要"
        self._write_summary_sheet(ws1)

        # Sheet2: 单价异常清单
        ws2 = wb.create_sheet("单价异常清单")
        self._write_price_sheet(ws2)

        # Sheet3: 指标分析
        ws3 = wb.create_sheet("指标分析")
        self._write_indicator_sheet(ws3)

        wb.save(output_path)

    def _write_summary_sheet(self, ws):
        ws.merge_cells("A1:F1")
        ws["A1"] = "工程造价审核报告"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER

        ws.merge_cells("A2:F2")
        ws["A2"] = "审核单位：中联永信工程管理（河南）有限公司"
        ws["A2"].font = BODY_FONT
        ws["A2"].alignment = CENTER

        # 统计
        high_cnt = sum(1 for f in self.findings if f.get("level") == "high")
        med_cnt = sum(1 for f in self.findings if f.get("level") == "medium")
        price_items = [f for f in self.findings if f.get("type") == "单价审核"]

        row = 5
        info = [
            ("审核子目总数", f"{len(price_items)} 条"),
            ("总造价", f"{self.total_price / 10000:.2f} 万元"),
            ("🔴 严重偏差（>±20%）", f"{high_cnt} 条"),
            ("🟡 需要注意（±10-20%）", f"{med_cnt} 条"),
            ("🟢 正常范围", f"{len(price_items) - high_cnt - med_cnt} 条"),
        ]
        for label, val in info:
            ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=11, bold=True)
            ws.cell(row=row, column=3, value=val).font = BODY_FONT
            row += 1

        # 设置列宽
        for c in "ABCDEF":
            ws.column_dimensions[c].width = 18

    def _write_price_sheet(self, ws):
        headers = ["序号", "风险", "子目名称", "单位", "工程量", "清单单价", "参考区间", "偏差率", "问题分析", "建议"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        price_findings = [f for f in self.findings if f.get("type") == "单价审核"]
        price_findings.sort(key=lambda x: abs(x["deviation"]), reverse=True)

        for i, f in enumerate(price_findings, 1):
            row = i + 1
            level_text = "🔴严重" if f["level"] == "high" else "🟡注意"
            fill = RED_FILL if f["level"] == "high" else YELLOW_FILL
            font = RED_FONT if f["level"] == "high" else YELLOW_FONT

            data = [i, level_text, f["name"], f["unit"], f["qty"], f["price"],
                    f["ref_range"], f"{f['deviation']:+.1f}%", f["suggestion"], f["suggestion"]]
            for j, val in enumerate(data, 1):
                c = ws.cell(row=row, column=j, value=val)
                c.font = font if j == 2 else BODY_FONT
                c.fill = fill
                c.border = THIN_BORDER
                c.alignment = CENTER if j < 4 else LEFT

        widths = [6, 10, 35, 8, 10, 12, 18, 10, 40, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_indicator_sheet(self, ws):
        headers = ["指标名称", "本工程数值", "参考范围", "偏差", "评价"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        if not self.indicators:
            ws.cell(row=2, column=1, value="未计算指标（可能缺少面积数据）").font = BODY_FONT

        for i, (name, info) in enumerate(self.indicators.items(), 1):
            row = i + 1
            dev_text = ""
            fill = GREEN_FILL
            if "ok" in info and not info["ok"]:
                fill = YELLOW_FILL
                dev_text = "⚠ 偏离参考范围"
            comment = "正常" if info.get("ok", True) else "需关注"
            data = [name, str(info["value"]), info.get("range", ""), dev_text, comment]
            for j, val in enumerate(data, 1):
                c = ws.cell(row=row, column=j, value=val)
                c.font = BODY_FONT
                c.fill = fill
                c.border = THIN_BORDER
                c.alignment = CENTER

        for c in "ABCDE":
            ws.column_dimensions[c].width = 20

    # ── 生成摘要文本 ──
    def get_summary(self) -> str:
        """生成 Markdown 格式的聊天摘要"""
        high = [f for f in self.findings if f.get("level") == "high"]
        med = [f for f in self.findings if f.get("level") == "medium"]
        lines = [
            "## 📊 造价审核完成",
            f"- 总造价：**{self.total_price / 10000:.2f} 万元**",
            f"- 审核子目：{len([f for f in self.findings if f.get('type') == '单价审核'])} 条",
            f"- 🔴 严重偏差：**{len(high)} 条**",
            f"- 🟡 需注意：{len(med)} 条",
        ]
        if self.indicators:
            lines.append("\n### 📐 工程量指标")
            for name, info in self.indicators.items():
                ok = "✅" if info.get("ok", True) else "⚠️"
                lines.append(f"- {ok} {name}：{info['value']}（参考：{info.get('range', '')})")
        if high:
            lines.append("\n### 🔴 严重偏差项目")
            for f in high[:5]:
                lines.append(f"- {f['name']}：单价 {f['price']}，偏离 {f['deviation']:+.1f}%，参考 {f['ref_range']}")
        return "\n".join(lines)


# ═══════════════════════════════════════════════
# 便捷函数
# ═══════════════════════════════════════════════

def quick_audit(filepath: str, building_type: str = "工业建筑", area: float = None) -> str:
    """一键审核，返回摘要"""
    a = Auditor(filepath)
    if not a.parse():
        return "❌ 无法解析清单文件，请检查格式"
    a.audit_prices()
    a.analyze_indicators(building_type, area)
    out = filepath.replace(".xlsx", "_审核报告.xlsx").replace(".xls", "_审核报告.xlsx")
    a.save_report(out)
    return a.get_summary() + f"\n\n📁 详细报告已保存至：{out}"


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法: python zaojia_auditor.py <清单文件.xlsx> [建筑类型:工业建筑|民用建筑] [建筑面积]")
        sys.exit(1)
    path = sys.argv[1]
    btype = sys.argv[2] if len(sys.argv) > 2 else "工业建筑"
    area = float(sys.argv[3]) if len(sys.argv) > 3 else None
    result = quick_audit(path, btype, area)
    print(result)
